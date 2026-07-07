"""Bob file ingestion — extract text from uploaded resumes / cohort sheets.

Supported: PDF (PyMuPDF), DOCX (zip + document.xml, same approach as the
candidate_intelligence parser), XLSX (openpyxl), CSV/TXT (plain decode).
Extraction happens at upload time; the extracted text is stored on bob_files
and injected into the agent's context on every run in that chat.
"""

import io
import logging
import re
import xml.etree.ElementTree as ET
import zipfile

logger = logging.getLogger(__name__)

MAX_TEXT_CHARS = 20000  # per file, at ingest
MAX_XLSX_ROWS = 300
MAX_XLSX_SHEETS = 5


class FileExtractionError(Exception):
    pass


def extract_text(filename: str, data: bytes) -> str:
    name = (filename or "").lower()
    try:
        if name.endswith(".pdf"):
            text = _pdf_text(data)
        elif name.endswith(".docx"):
            text = _docx_text(data)
        elif name.endswith((".xlsx", ".xlsm")):
            text = _xlsx_text(data)
        elif name.endswith((".csv", ".txt")):
            text = data.decode("utf-8", errors="replace")
        else:
            raise FileExtractionError(
                "Unsupported file type. Supported: PDF, DOCX, XLSX, CSV, TXT."
            )
    except FileExtractionError:
        raise
    except Exception as e:  # noqa: BLE001
        logger.warning("[BOB/FILES] extraction failed for %s: %s", filename, e)
        raise FileExtractionError(f"Could not read {filename}: {e}") from e

    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    if not text:
        raise FileExtractionError(
            f"No text could be extracted from {filename}. If it is a scanned PDF, export it as a text PDF first."
        )
    return text[:MAX_TEXT_CHARS]


def _pdf_text(data: bytes) -> str:
    import fitz  # PyMuPDF

    doc = fitz.open(stream=data, filetype="pdf")
    try:
        return "\n".join(page.get_text() for page in doc[:20])
    finally:
        doc.close()


def _docx_text(data: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        xml_data = z.read("word/document.xml")
    root = ET.fromstring(xml_data)
    ns = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    parts = []
    for para in root.iter(f"{ns}p"):
        runs = [node.text or "" for node in para.iter(f"{ns}t")]
        if runs:
            parts.append("".join(runs))
    return "\n".join(parts)


def _xlsx_text(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    for ws in wb.worksheets[:MAX_XLSX_SHEETS]:
        parts.append(f"### Sheet: {ws.title}")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_XLSX_ROWS:
                parts.append(f"... (truncated at {MAX_XLSX_ROWS} rows)")
                break
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                parts.append(" | ".join(cells))
    wb.close()
    return "\n".join(parts)
